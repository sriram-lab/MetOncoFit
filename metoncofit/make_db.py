#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
db.py creates the metoncofit dataframe that can be used for several web and development applications.
@author: Scott Campit
"""

import sys
import copy
import operator
import argparse
import re
import os

import numpy as np
import pandas as pd

from sklearn import preprocessing
from sklearn.preprocessing import RobustScaler
from sklearn.preprocessing import MinMaxScaler
from imblearn.over_sampling import RandomOverSampler
from sklearn.model_selection import train_test_split
from sklearn.externals import joblib
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import cross_val_score

datapath = None
all_dfs = []
targ = ["TCGA_annot", "CNV", "SURV"]
var_excl = ["TCGA gene expression fold change", "CNV gain/loss ratio"]

x = 0
for fil in os.listdir('./../data/median/'):
    # Iterate between models
    x = x+1
    for t in targ:

        # Proprocessing data
        classes = []
        data = []
        names = []

        if t == 'TCGA_annot':
            t = str("TCGA annotation")
        if datapath is None:
            datapath = './../data/original/'

        canc = fil.replace(".csv", "")
        canc_dict = {
            'breast': 'Breast Cancer',
            'cns': 'Glioma',
            'colon': 'Colorectal Cancer',
            'complex': 'Pan Cancer',
            'leukemia': 'B-cell lymphoma',
            'melanoma': 'Melanoma',
            'nsclc': 'Lung Cancer',
            'ovarian': 'Ovarian Cancer',
            'prostate': 'Prostate Cancer',
            'renal': 'Renal Cancer'
            }
        canc = canc_dict.get(canc)

        df_names = pd.read_csv(
            "./../labels/real_headers.txt", sep='\t', names=['Original', 'New'])
        names = dict([(i, nam)
                      for i, nam in zip(df_names['Original'], df_names['New'])])
        df = pd.read_csv(datapath+fil, index_col=None)
        df = df.drop(columns=['TCGA_val', 'CNV_val'], axis=1)
        df = df.rename(columns=names)
        df = df.set_index(['Genes', 'Cell Line'])

        # We are label encoding the subsystem and datapath labels
        le = preprocessing.LabelEncoder()
        df["RECON1 subsystem"] = le.fit_transform(df["RECON1 subsystem"])
        df["Metabolic subnetwork"] = le.fit_transform(
            df["Metabolic subnetwork"])

        excl_targ = {'TCGA annotation', 'SURV', 'CNV'}
        tmp = excl_targ.remove(t)

        df = df.drop(columns=excl_targ)
        classes = df[t]
        header = df.columns
        df1 = df.copy(deep=True)  # contains target classes
        df = df.drop(columns=t)  # doesn't contain target classes

        data = np.array(df).astype(np.float)
        data = RobustScaler().fit_transform(data)

        new_data, orig_data, new_classes, orig_classes = train_test_split(
            data, classes, test_size=0.3)

        ros = RandomOverSampler()
        data, classes = ros.fit_sample(new_data, new_classes)

        # Random forests (MetOncoFit)
        feat = (data.shape[1]-10)
        while(feat < data.shape[1]-1):
            trees = 5
            while(trees <= 500):
                rfc = RandomForestClassifier(
                    n_estimators=trees, max_features=feat)
                rfc.fit(data, classes)
                trees = trees + 1500
            feat = feat + 20
            rfc_pred = rfc.predict(orig_data)
            mean_acc = rfc.score(orig_data, orig_classes)

        if(t == "CNV"):
            targ_labels = ["GAIN", "NEUT", "LOSS"]
            targ_dict = {'NEUT': 0, 'LOSS': 0, 'GAIN': 0}
        else:
            targ_labels = ["UPREG", "NEUTRAL", "DOWNREG"]
            targ_dict = {'NEUTRAL': 0, 'DOWNREG': 0, 'UPREG': 0}

        df1 = df1.reset_index()
        one_gene_df = df1.drop(columns="Cell Line").groupby(
            ["Genes", t]).median().reset_index().set_index("Genes")
        one_gene_class = pd.DataFrame(one_gene_df[t])
        one_gene_class = one_gene_class.reset_index()

        # These dataframes contain the df entries with increased, neutral, and decreased values.
        up_df = one_gene_df.loc[(one_gene_df[t] == targ_labels[0])]
        neut_df = one_gene_df.loc[(one_gene_df[t] == targ_labels[1])]
        down_df = one_gene_df.loc[(one_gene_df[t] == targ_labels[2])]

        # To create the figure, we are randomly selecting three genes that are upreg, neutral, or downreg and are storing them in this list.
        up_genes = up_df.index.values.tolist()
        neut_genes = neut_df.index.values.tolist()
        down_genes = down_df.index.values.tolist()

        # Remove the classes
        _ = one_gene_df.pop(t)

        # This will calculate the correlation for each feature, if there is one between the biological features.
        column_squigly = {}
        for col in one_gene_df.columns:
            v1 = up_df[col].median()
            v2 = neut_df[col].median()
            v3 = down_df[col].median()

            correl = np.corrcoef([v1, v2, v3], [1.0, 0.0, -1.0])

            if(np.isnan(correl[0][1]) != True):
                column_squigly[col] = correl[0][1]
            else:
                column_squigly[col] = 0.0

        def idx_change(header, to_be_mapped):
            """
            idx_change sorts the feature importances and maps it to the feature name
            """
            temp_dict_feat = {}
            for i, j in zip(header, to_be_mapped):
                temp_dict_feat[i] = j
            sorted_d = sorted(temp_dict_feat.items(),
                              key=operator.itemgetter(1), reverse=True)
            return sorted_d
        sorted_d = idx_change(header, rfc.feature_importances_)

        feat = []
        gini = []
        corr = []

        x = 0
        while(x < 130):  # Get the first 10 features
            tempa = sorted_d[x]
            feat.append(tempa[0])
            gini.append(tempa[1])
            corr.append((column_squigly[tempa[0]]))
            x = x+1

        importance = pd.DataFrame({"Feature": feat, "Gini": gini, "R": corr})

        # Map to label
        if t == 'CNV':
            class_col = ["GAIN", "NEUT", "LOSS"]
        else:
            class_col = ["UPREGULATED", "NEUTRAL", "DOWNREGULATED"]

        # Scale the dataframe from 0 to 1
        idx = one_gene_df.index
        col = one_gene_df.columns
        scaler = MinMaxScaler()
        result = scaler.fit_transform(one_gene_df)
        one_gene_df = pd.DataFrame(result, columns=col, index=idx)

        # Get the genes that are up/neut/downregulated
        tmparr_up = one_gene_df[one_gene_df.index.isin(up_genes)]
        tmparr_neut = one_gene_df[one_gene_df.index.isin(neut_genes)]
        tmparr_down = one_gene_df[one_gene_df.index.isin(down_genes)]

        features = list(importance['Feature'])

        up = tmparr_up[features].T
        up = up.merge(importance, how='inner',
                      left_index=True, right_on='Feature')
        up = pd.melt(up, id_vars=["Feature", "Gini", "R"],
                     var_name="Gene", value_name="Value")
        up["Type"] = class_col[0]

        neut = tmparr_neut[features].T
        neut = neut.merge(importance, how='inner',
                          left_index=True, right_on='Feature')
        neut = pd.melt(
            neut, id_vars=["Feature", "Gini", "R"], var_name="Gene", value_name="Value")
        neut["Type"] = class_col[1]

        down = tmparr_down[features].T
        down = down.merge(importance, how='inner',
                          left_index=True, right_on='Feature')
        down = pd.melt(
            down, id_vars=["Feature", "Gini", "R"], var_name="Gene", value_name="Value")
        down["Type"] = class_col[2]

        final_df = pd.concat([up, neut, down], axis=0)
        final_df['Cancer'] = canc
        final_df['Cancer'] = final_df['Cancer'].str.replace(' Cancer', '')
        final_df = final_df.reset_index().drop('index', axis=1)

        if t == "TCGA annotation":
            t = "Differential Expression"
        elif t == "CNV":
            t = "Copy Number Variation"
        else:
            t = "Patient Survival"

        final_df["Target"] = t
        final_df = final_df[~final_df['Cancer'].str.contains('Pan')]

        # Improve data storage efficiency
        final_df = final_df.round(2)
        final_df['Gini'] = pd.to_numeric(final_df['Gini'], downcast='float')
        final_df['R'] = pd.to_numeric(final_df['R'], downcast='float')
        final_df['Value'] = pd.to_numeric(final_df['Value'], downcast='float')
        final_df['Type'] = final_df['Type'].astype('category')
        final_df['Cancer'] = final_df['Cancer'].astype('category')
        final_df['Target'] = final_df['Target'].astype('category')
        final_df['Feature'] = final_df['Feature'].astype('category')
        final_df['Gene'] = final_df['Gene'].astype('category')
        canc = canc.replace(' Cancer', '')
        print(canc)

        from openpyxl import load_workbook
        book = load_workbook('./../output/Tables/SI.xlsx')
        writer = pd.ExcelWriter(
            './../output/Tables/SI.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        if ("S. Table "+str(x+9)+" | "+canc) in book:
            final_df.to_excel(writer, sheet_name=("S. Table "+str(x+9)+" | "+canc),
                              startrow=writer.sheets[("S. Table "+str(x+9)+" | "+canc)].max_row, header=False)
        else:
            final_df.to_excel(writer, sheet_name=(
                "S. Table "+str(x+9)+" | "+canc), header=True)
            writer.save()

#big_df = pd.concat(all_dfs, axis=0, ignore_index=True)
#big_df.to_csv("db.csv")
#big_df.to_json("db.json")
