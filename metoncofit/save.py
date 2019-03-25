#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
`save.py` saves all the resulting data into Excel files
@authors: Krishna Oruganty & Scott Campit
"""

import sys
import pandas as pd
from openpyxl import load_workbook

targ = sys.argv[2]

def make_excel(summary, compare_models, loco, lofo):
    """
    make_excel is the function that will save the resulting dataframes into Excel Files.

    INPUTS:
        * summary: The summary statistics for predicting differential expression, copy number variation, and cancer patient survival.
        * compare_models: The AUROC scores that will be used to compare MetOncoFit against MCT-SVM.
        * loco: The leave-one-cell-line-out accuracy scores that will assess how each cell line contributes to the model.
        * lofo: The leave-one-feature-out accuracy scores that will assess how each feature set contributes to predicting the prognostic cancer targets.

    OUTPUT:
        * An excel file in the output/Tables folder containing the Excel file.
    """

    book = load_workbook('./../output/Tables/SI.xlsx')
    writer = pd.ExcelWriter('./../output/Tables/SI.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Save the summary statistics data into excel files
    if targ == "CNV":
        if "S. Table 4 | CNV Pred" in book:
            summary.to_excel(writer, sheet_name="S. Table 4 | CNV Pred", startrow=writer.sheets["S. Table 4 | CNV Pred"].max_row, header=False)
        else:
            summary.to_excel(writer, sheet_name="S. Table 4 | CNV Pred", header=True)

    elif targ == "SURV":
        if "S. Table 5 | SURV Pred" in book:
            summary.to_excel(writer, sheet_name="S. Table 5 | SURV Pred", startrow=writer.sheets["S. Table 5 | SURV Pred"].max_row, header=False)
        else:
            summary.to_excel(writer, sheet_name="S. Table 5 | SURV Pred", header=True)

    elif targ == "TCGA annotation":
        if "S. Table 3 | DE Pred" in book:
            summary.to_excel(writer, sheet_name="S. Table 3 | DE Pred", startrow=writer.sheets["S. Table 3 | DE Pred"].max_row, header=False)
        else:
            summary.to_excel(writer, sheet_name="S. Table 3 | DE Pred", header=True)

    # Save the Leave-One-Feature-Out dataset
    if "S. Table 6 | LOFO" in book:
        lofo.to_excel(writer, sheet_name="S. Table 6 | LOFO", index=False, startrow=writer.sheets["S. Table 6 | LOFO"].max_row, header=False)
    else:
        lofo.to_excel(writer, sheet_name="S. Table 6 | LOFO", index=False, header=True)

    # Save the Leave-One-Cell-Out dataset
    if "S. Table 7 | LOCO" in book:
        loco.to_excel(writer, sheet_name="S. Table 7 | LOCO", index=False, startrow=writer.sheets["S. Table 7 | LOCO"].max_row, header=False)
    else:
        loco.to_excel(writer, sheet_name="S. Table 7 | LOCO", index=False, header=True)

    # Save AUROC scores
    if "S. Table 8 | AUROC" in book:
        compare_models.to_excel(writer, sheet_name="S. Table 8 | AUROC", index=False, startrow=writer.sheets["S. Table 8 | AUROC"].max_row, header=False)
    else:
        compare_models.to_excel(writer, sheet_name="S. Table 8 | AUROC", index=False, header=True)

    writer.save()
