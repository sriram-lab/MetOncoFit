"""
HR Check
"""
import sys
import numpy as np
import scipy
import pandas as pd
from openpyxl import load_workbook

import process
import random_forest
import validator
import visualizations
import save

# Get the frequency for each label in each dataset
df, df1, header, canc, targ, data, classes, orig_data, orig_classes, excl_targ, freq = process.preprocess(datapath=sys.argv[1], fil=sys.argv[2], targ=sys.argv[3], exclude=sys.argv[4])

# Random Forest
rfc, rfc_pred, mean_acc = random_forest.random_forest(canc, targ, data, classes, orig_data, orig_classes)

# Summary statistics
cm, pvalue, zscore, cv_score, summary = validator.summary_statistics(rfc, rfc_pred, data, classes, orig_classes, orig_data, targ, excl_targ, mean_acc, canc)

freq["10-fold CV Accuracy"] = cv_score

# Save the results in an Excel file
book = load_workbook('./../output/Tables/SI.xlsx')
writer = pd.ExcelWriter(r'./../output/Tables/SI.xlsx', engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
if "S. Table 9 | HR Check" in book:
    freq.to_excel(writer, sheet_name="S. Table 9 | HR Check", startrow=writer.sheets["S. Table 9 | HR Check"].max_row, header=False)
else:
    freq.to_excel(writer, sheet_name="S. Table 9 | HR Check", header=True)
writer.save()
